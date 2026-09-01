import openpyxl
import re

from datetime import datetime, time
from record_classifier import classify_record

def parse_time(value):
    # Empty Excel cell
    if value is None:
        return None

    # Excel stored the value as a time object
    if isinstance(value, time):
        return value

    # Convert the value to text
    value = str(value).strip()

    # Remove manual-correction marker, if present
    value = value.replace("*", "").strip()

    # Convert HH:MM text into a datetime.time object
    return datetime.strptime(value, "%H:%M").time()

def read_attendance(filename, employment_end_date):

    # Open the Excel file in read-only mode
    workbook = openpyxl.load_workbook(filename, read_only=True)

    # Select the first worksheet
    sheet = workbook.active

    # Extract the reporting period from row 2
    period = next(
        sheet.iter_rows(
            min_row=2,
            max_row=2,
            values_only=True
        )
    )[0]

    # Split the reporting period into start and end dates
    start_text, end_text = period.split(" - ")

    # Convert the text dates into Python date objects
    start_date = datetime.strptime(
        start_text,
        "%d.%m.%Y"
    ).date()

    end_date = datetime.strptime(
        end_text,
        "%d.%m.%Y"
    ).date()

    # Store extracted attendance records here
    records = []

    # Store the percentage-based salary columns found in the report
    percentage_columns = []

    # Read the attendance table starting from row 3
    for row in sheet.iter_rows(min_row=3, values_only=True):

        # Column A contains the date label
        date_text = row[0]

        # Empty row means the useful table ended
        if date_text is None:
            break

        # Convert to string for processing
        date_text = str(date_text)

        # Keep the header row
        if date_text == "תאריך":

            # Find all columns whose header contains "%"
            for index, column_name in enumerate(row):
                if column_name is not None and "%" in str(column_name):
                    percentage_columns.append({
                        "name": column_name,
                        "index": index
                    })

            # Find the event column
            event_column = None

            for index, column_name in enumerate(row):
                if column_name is not None and str(column_name).strip() == "אירוע":
                    event_column = index

            records.append({
                "date": "תאריך",
                "weekday_he": None,
                "type": row[1],
                "entry": row[2],
                "exit": row[3],
                "event": row[event_column]
            })

            continue

        # Ignore weekly/monthly summary rows
        if not re.fullmatch(r"[אבגדהוש] - \d{2}", date_text):
            continue

        # Extract the day number
        day = int(date_text.split("-")[1])

        # Build the complete date
        actual_date = start_date.replace(day=day)

        # Ignore attendance after the employment end date
        if actual_date > employment_end_date:
            continue

        # Keep the original Hebrew weekday
        weekday_he = date_text.split("-")[0].strip()

        # Normalize the entry and exit times
        entry = parse_time(row[2])
        exit = parse_time(row[3])

        event = row[event_column]

        # strip spaces from string
        if event is not None:
            event = str(event).strip()

        # Extract the percentage-based salary components
        percentages = {}

        for column in percentage_columns:
            percentages[column["name"]] = row[column["index"]]

        # Store the relevant information
        record = {
            "date": actual_date,
            "weekday_he": weekday_he,
            "type": row[1],
            "entry": entry,
            "exit": exit,
            "event": event,
            "percentages": percentages
        }

        # Classify the record based on its type, event, and attendance times
        record["category"] = classify_record(record)

        # Add the record to the list
        records.append(record)

    # Return the extracted records to the caller
    return records
