import openpyxl
import os

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from calc_base_salary import get_hourly_rate, calculate_base_salary
from add_sum_row import add_sum_row

def format_time(value):
    # Keep empty cells empty
    if value is None:
        return ""

    # Format time as HH:MM
    return value.strftime("%H:%M")


def format_duration(value):
    # Keep empty cells empty
    if value is None:
        return ""

    # Convert the duration into total minutes
    total_minutes = int(value.total_seconds() // 60)

    # Calculate hours and remaining minutes
    hours = total_minutes // 60
    minutes = total_minutes % 60

    # Always display two digits for hours and minutes
    return f"{hours:02d}:{minutes:02d}"


CATEGORY_NAMES = {
    "WORK_DAY": "יום עבודה",
    "SICKNESS": "מחלה",
    "VACATION": "חופשה",
    "HOLIDAY": "חג",
    "WEEKEND": "סוף שבוע",
    "OTHER": "אחר"
}


def export_xlsx(records, input_filename):

    # Extract the filename without its directory
    filename = os.path.basename(input_filename)

    # Build the output filename
    output_filename = os.path.join(
        "output",
        os.path.splitext(filename)[0] + ".xlsx"
    )

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    sheet = workbook.active

    # Give the worksheet a meaningful name
    sheet.title = "נוכחות"

    # Display the worksheet from right to left
    sheet.sheet_view.rightToLeft = True

    # Keep track of the previous date
    previous_date = None

    # Get the percentage columns from the first actual record
    percentage_names = []

    for record in records:
        if record["date"] != "תאריך":
            percentage_names = list(record["percentages"].keys())
            break

    # Write the records to the worksheet
    for record in records:

        # Handle the header row
        if record["date"] == "תאריך":

            sheet.append([
                record["date"],
                "יום",
                record["type"],
                record["entry"],
                record["exit"],
                "קטגוריה",
                "כניסת שכר יסוד",
                "יציאת שכר יסוד",
                "שעות שכר יסוד",
                "שעות שכר יסוד (עשרוני)",
                *percentage_names,
                "שכר לשעה",
                "תוספת לילה 25%",
                "שכר יסוד",
            ])

            # Make the header bold
            for cell in sheet[sheet.max_row]:
                cell.font = Font(bold=True)

            # Give the header enough height for wrapped text
            sheet.row_dimensions[1].height = 38

            continue

        # Determine the hourly salary rate
        hourly_rate = get_hourly_rate(record["date"])

        # get base salary
        salary_calculation = calculate_base_salary(
            record,
            hourly_rate
        )

        base_salary = salary_calculation["base_salary"]
        night_premium = salary_calculation["night_premium"]

        # Add an empty row when a new week begins
        if previous_date is not None:
            if record["weekday_he"] == "א":
                sheet.append([])

        # Write the attendance record
        sheet.append([
            record["date"],
            record["weekday_he"],
            record["type"],
            format_time(record["entry"]),
            format_time(record["exit"]),
            CATEGORY_NAMES[record["category"]],
            format_time(record.get("base_entry")),
            format_time(record.get("base_exit")),
            (
                record["base_duration"].total_seconds() / 86400
                if record.get("base_duration") is not None
                else None
            ),
            record.get("base_hours"),
            *[
                record["percentages"].get(name)
                for name in percentage_names
            ],
            hourly_rate,
            night_premium if night_premium != 0 else None,
            base_salary
        ])

        # Get the row that was just written
        current_row = sheet.max_row

        # Format the decimal base-hours cell as a number
        # with exactly two decimal places.
        #
        # The underlying value remains a real number,
        # so Excel can use it in SUM() and other calculations.

        # Format base duration as an Excel duration.
        # [h]:mm allows totals greater than 24 hours.
        sheet.cell(
            row=current_row,
            column=9
        ).number_format = "[hh]:mm"

        sheet.cell(
            row=current_row,
            column=10
        ).number_format = "0.00"

        # Format hourly rate as a number
        sheet.cell(
            row=current_row,
            column=11 + len(percentage_names)
        ).number_format = "0.00"

        # format the two new monetary columns
        sheet.cell(
            row=current_row,
            column=12 + len(percentage_names)
        ).number_format = "0.00"

        sheet.cell(
            row=current_row,
            column=13 + len(percentage_names)
        ).number_format = "0.00"

        # Remember this record's date
        previous_date = record["date"]

    # Adjust column widths for readability
    column_widths = {
        1: 15,  # תאריך
        2: 8,   # יום
        3: 12,  # סוג
        4: 10,  # כניסה
        5: 10,  # יציאה
        6: 14,  # קטגוריה
        7: 18,  # כניסת שכר יסוד
        8: 18,  # יציאת שכר יסוד
        9: 16,  # שעות שכר יסוד
        10: 24,  # שעות שכר יסוד (עשרוני)
        11 + len(percentage_names): 12,  # שכר לשעה
        12 + len(percentage_names): 18,  # תוספת לילה 25%
        13 + len(percentage_names): 14   # שכר יסוד
    }

    for column, width in column_widths.items():
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = width

    # Add monthly summary row
    add_sum_row(sheet)

    # Center-align the entire table
    # Keep the header wrapped so long column names fit.
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=(cell.row == 1)
            )

    # Save the workbook
    workbook.save(output_filename)

    # Return the generated filename
    return output_filename
