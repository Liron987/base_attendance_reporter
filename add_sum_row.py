from openpyxl.styles import Font


def add_sum_row(sheet):

    # Add the summary after the last existing row
    summary_row = sheet.max_row + 1

    # Write the summary label
    sheet.cell(
        row=summary_row,
        column=1
    ).value = "סיכום חודשי"

    # Find columns by their header names
    headers = {
        cell.value: cell.column
        for cell in sheet[1]
        if cell.value is not None
    }

    # Sum the base-hours duration column
    duration_column = headers.get("שעות שכר יסוד")

    if duration_column is not None:
        sheet.cell(
            row=summary_row,
            column=duration_column
        ).value = (
            f"=SUM("
            f"{sheet.cell(row=2, column=duration_column).coordinate}:"
            f"{sheet.cell(row=summary_row - 1, column=duration_column).coordinate}"
            f")"
        )

        sheet.cell(
            row=summary_row,
            column=duration_column
        ).number_format = "[h]:mm"

    # Sum the decimal base-hours column
    decimal_column = headers.get("שעות שכר יסוד (עשרוני)")

    if decimal_column is not None:
        sheet.cell(
            row=summary_row,
            column=decimal_column
        ).value = (
            f"=SUM("
            f"{sheet.cell(row=2, column=decimal_column).coordinate}:"
            f"{sheet.cell(row=summary_row - 1, column=decimal_column).coordinate}"
            f")"
        )

        sheet.cell(
            row=summary_row,
            column=decimal_column
        ).number_format = "0.00"

    # Sum the night premium
    night_column = headers.get("תוספת לילה 25%")

    if night_column is not None:
        sheet.cell(
            row=summary_row,
            column=night_column
        ).value = (
            f"=SUM("
            f"{sheet.cell(row=2, column=night_column).coordinate}:"
            f"{sheet.cell(row=summary_row - 1, column=night_column).coordinate}"
            f")"
        )

        sheet.cell(
            row=summary_row,
            column=night_column
        ).number_format = "0.00"

    # Sum the base salary
    salary_column = headers.get("שכר יסוד")

    if salary_column is not None:
        sheet.cell(
            row=summary_row,
            column=salary_column
        ).value = (
            f"=SUM("
            f"{sheet.cell(row=2, column=salary_column).coordinate}:"
            f"{sheet.cell(row=summary_row - 1, column=salary_column).coordinate}"
            f")"
        )

        sheet.cell(
            row=summary_row,
            column=salary_column
        ).number_format = "0.00"

    # Make the summary row bold
    for cell in sheet[summary_row]:
        cell.font = Font(bold=True)
